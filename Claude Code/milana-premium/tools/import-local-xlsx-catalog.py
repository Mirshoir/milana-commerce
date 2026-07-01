#!/usr/bin/env python3
"""Import local Milana XLSX catalog rows and crop product card images from PDFs."""

from __future__ import annotations

import json
import re
import sqlite3
import subprocess
from pathlib import Path

import openpyxl
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
TMP_DIR = ROOT / "tmp" / "pdfs" / "xlsx-import"
XLSX = Path("/Users/isomiddin/Downloads/milana_products_latest (1).xlsx")
PDFS = {
    "01_Staple_Model_Catalog.pdf": Path("/Users/isomiddin/Downloads/2026_Каталог_моделей_из_Штапеля_19_05_2026.pdf"),
    "02_Milana_Man_Premium_Collection.pdf": Path("/Users/isomiddin/Downloads/Milana Man Premium Collection 19.05.2026.pdf"),
    "03_Kindergarten_Set.pdf": Path("/Users/isomiddin/Downloads/Комплект для Садика 19.05.2026.pdf"),
}
META = {
    "01_Staple_Model_Catalog.pdf": {"gender": "women", "category": "loungewear"},
    "02_Milana_Man_Premium_Collection.pdf": {"gender": "men", "category": "loungewear"},
    "03_Kindergarten_Set.pdf": {"gender": "kids", "category": "pajamas"},
}
DEFAULT_SIZES = {
    "women": ["44", "46", "48", "50", "52", "54"],
    "men": ["46", "48", "50", "52", "54", "56"],
    "kids": ["28", "30", "32", "34", "36", "38"],
}
PDFTOPPM = Path("/Users/isomiddin/.cache/codex-runtimes/codex-primary-runtime/dependencies/bin/pdftoppm")


def slugify(value: str) -> str:
    value = value.lower()
    value = re.sub(r"['’`ʻ]", "", value)
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")[:80] or "item"


def clean_text(value, max_len: int = 5000) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text).strip()[:max_len]


def parse_sizes(text: str, gender: str) -> list[str]:
    found: list[str] = []
    seen: set[str] = set()
    for token in re.split(r"\s+", text or ""):
        clean = re.sub(r"[^\dA-Za-z]", "", token).upper()
        ok = clean.isdigit() and len(clean) == 2 and 24 <= int(clean) <= 60 and int(clean) % 2 == 0
        if ok and clean not in seen:
            seen.add(clean)
            found.append(clean)
    for size in DEFAULT_SIZES[gender]:
        if size not in seen:
            found.append(size)
            seen.add(size)
    return found[:6]


def fabric_text(row: dict) -> str:
    text = str(row.get("combined_text") or row.get("native_text") or row.get("ocr_text") or "")
    lines = []
    for line in text.splitlines():
        clean = line.strip()
        if not clean:
            continue
        if re.match(r"^(MIL[>A-Z]*|PREMIUM|MODEL|CODE|SALE)$", clean, re.I):
            continue
        if re.match(r"^\d+(\.\d+)?\s*\$?$", clean, re.I):
            continue
        if re.match(r"^[A-Z]-?\d+$", clean, re.I):
            continue
        if re.match(r"^\d{2}$", clean):
            continue
        lines.append(clean)
    return clean_text(" · ".join(lines[:2]), 300)


def product_name(row: dict) -> str:
    model = clean_text(row.get("model_code"), 80)
    code = clean_text(row.get("product_code"), 80)
    if model and code and model != code:
        return f"{model} / {code}"
    return model or code or f"Catalog item {row.get('page', '')}-{row.get('card_index', '')}"


def render_page(pdf: Path, source_pdf: str, page: int) -> Path:
    out_dir = TMP_DIR / slugify(source_pdf)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_prefix = out_dir / f"p{page:03d}"
    rendered = out_dir / f"p{page:03d}-{page:02d}.png"
    if rendered.exists():
        return rendered
    subprocess.run(
        [str(PDFTOPPM), "-r", "180", "-png", "-f", str(page), "-l", str(page), str(pdf), str(out_prefix)],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    matches = sorted(out_dir.glob(f"p{page:03d}-*.png"))
    if not matches:
        raise RuntimeError(f"Could not render {pdf} page {page}")
    return matches[0]


def crop_image(row: dict) -> str:
    source = row["source_pdf"]
    pdf = PDFS[source]
    page = int(row["page"])
    bbox = [int(float(x)) for x in str(row["bbox"]).split(",")]
    page_png = render_page(pdf, source, page)
    with Image.open(page_png) as image:
        x1, y1, x2, y2 = bbox
        x1 = max(0, min(x1, image.width - 1))
        y1 = max(0, min(y1, image.height - 1))
        x2 = max(x1 + 1, min(x2, image.width))
        y2 = max(y1 + 1, min(y2, image.height))
        card = image.crop((x1, y1, x2, y2)).convert("RGB")
        source_short = slugify(source.replace(".pdf", ""))
        name = f"catalog-{source_short}-p{page:03d}-c{int(row['card_index']):03d}.jpg"
        out = UPLOAD_DIR / name
        card.save(out, "JPEG", quality=88, optimize=True)
    return "/uploads/" + name


def read_rows() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Products"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if row.get("source_pdf") in PDFS:
            try:
                price = float(row.get("price") or 0)
            except (TypeError, ValueError):
                price = 0
            if price > 0 and row.get("bbox"):
                row["price"] = price
                rows.append(row)
    return rows


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing workbook: {XLSX}")
    missing = [str(p) for p in PDFS.values() if not p.exists()]
    if missing:
        raise SystemExit("Missing PDFs: " + ", ".join(missing))
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    TMP_DIR.mkdir(parents=True, exist_ok=True)

    rows = read_rows()
    db = sqlite3.connect(DATA_DIR / "milana.db")
    db.execute("PRAGMA busy_timeout = 5000")

    cols = [
        "slug", "model_no", "variant", "gender", "category", "name",
        "desc_en", "desc_ru", "desc_uz", "fabric_en", "fabric_ru", "fabric_uz",
        "price", "old_price", "sizes", "images", "tag", "rating", "reviews", "active", "sort", "created_at",
    ]
    placeholders = ",".join(["?"] * (1 + len(cols)))
    updates = ",".join(f"{c}=excluded.{c}" for c in cols)
    sql = f"""
      INSERT INTO products (slug, {",".join(cols)})
      VALUES ({placeholders})
      ON CONFLICT(slug) DO UPDATE SET {updates}
    """
    imported = 0
    by_source: dict[str, int] = {}
    try:
        db.execute("BEGIN")
        for row in rows:
            source = row["source_pdf"]
            meta = META[source]
            name = product_name(row)
            slug = "catalog-local-" + slugify(f"{source}-{row['page']}-{row['card_index']}-{name}")
            text = clean_text(row.get("combined_text") or row.get("native_text") or row.get("ocr_text"))
            image = crop_image(row)
            sizes = parse_sizes(text, meta["gender"])
            values = {
                "slug": slug,
                "model_no": clean_text(row.get("model_code"), 40),
                "variant": clean_text(row.get("product_code"), 60),
                "gender": meta["gender"],
                "category": meta["category"],
                "name": name,
                "desc_en": text,
                "desc_ru": text,
                "desc_uz": text,
                "fabric_en": fabric_text(row),
                "fabric_ru": fabric_text(row),
                "fabric_uz": fabric_text(row),
                "price": row["price"],
                "old_price": None,
                "sizes": json.dumps(sizes),
                "images": json.dumps([image]),
                "tag": "",
                "rating": 4.8,
                "reviews": 0,
                "active": 1,
                "sort": 1_100_000 - list(PDFS).index(source) * 10_000 - int(row["page"]) * 100 - int(row["card_index"]),
                "created_at": str(row.get("catalog_date") or "2026-05-25"),
            }
            db.execute(sql, [values["slug"]] + [values[c] for c in cols])
            imported += 1
            by_source[source] = by_source.get(source, 0) + 1
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        active = db.execute("SELECT gender, category, COUNT(*) FROM products WHERE active=1 GROUP BY gender, category").fetchall()
        total_active = db.execute("SELECT COUNT(*) FROM products WHERE active=1").fetchone()[0]
        db.close()

    print(json.dumps({
        "imported_or_updated": imported,
        "by_source": by_source,
        "active_products": total_active,
        "active_breakdown": active,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
